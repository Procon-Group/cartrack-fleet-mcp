#!/usr/bin/env python3
"""
Extracts the Procon Electrical & Steel Fleet Fuel Tracking System workbook into
dashboard/existing-fleet-data.json, matching the exact schema already baked into the
dashboard template (vehicles/order/months/monthlyLitres/monthlyCost/monthlyOther/
expenses/fuel/expenseTypeTotals). Read-only against the workbook - never writes to it.

Run this each time the user provides an updated workbook (a new month's fleet-card
statement merged in via the fleet-cost-workbook skill) - then run
`npm run dashboard:generate && npm run dashboard:build` and republish, since the Cost/KM,
Fuel Efficiency and Cost to Company target month all derive from this file's `months`/
`fuel` data.

Assumes the workbook's current layout: Fuel Log header row 5 (data from row 6), Vehicle
Expenses header row 5 (data from row 6), Fleet Register header row 4 (data from row 5).
Re-check these row numbers against references/workbook-structure.md (fleet-cost-workbook
skill) if the workbook's own structure has changed since - the workbook always wins.

Usage: python3 extract-workbook.py <path-to-xlsx> <output-json-path>
"""
import sys
import json
import openpyxl
from collections import defaultdict

def month_key(dt):
    return dt.strftime("%Y-%m") if dt else None

def date_key(dt):
    return dt.strftime("%Y-%m-%d") if dt else None

def main():
    xlsx_path, out_path = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---- Fuel Log (header row 5, data from row 6) ----
    ws = wb["Fuel Log"]
    fuel = []
    for r in range(6, ws.max_row + 1):
        reg = ws.cell(row=r, column=2).value
        if not reg:
            continue
        date = ws.cell(row=r, column=1).value
        fuel.append({
            "d": date_key(date),
            "r": reg,
            "dr": ws.cell(row=r, column=3).value or "Unassigned",
            "l": ws.cell(row=r, column=4).value or 0,
            "c": ws.cell(row=r, column=5).value or 0,
            "o": ws.cell(row=r, column=6).value,
            "v": ws.cell(row=r, column=8).value or "",
            "dv": ws.cell(row=r, column=9).value or "Unassigned",
            "p": ws.cell(row=r, column=10).value,
            "m": month_key(ws.cell(row=r, column=11).value),
            "s": ws.cell(row=r, column=12).value or "",
            "n": ws.cell(row=r, column=13).value or "",
        })

    # ---- Vehicle Expenses (header row 5, data from row 6) ----
    ws = wb["Vehicle Expenses"]
    expenses = []
    for r in range(6, ws.max_row + 1):
        reg = ws.cell(row=r, column=2).value
        if not reg:
            continue
        date = ws.cell(row=r, column=1).value
        expenses.append({
            "date": date_key(date),
            "reg": reg,
            "type": ws.cell(row=r, column=3).value or "",
            "desc": ws.cell(row=r, column=4).value or "",
            "cost": ws.cell(row=r, column=5).value or 0,
            "supplier": ws.cell(row=r, column=6).value or "",
            "vehicle": ws.cell(row=r, column=7).value or "",
            "division": ws.cell(row=r, column=8).value or "Unassigned",
            "month": month_key(ws.cell(row=r, column=9).value),
        })

    # ---- Fleet Register (header row 4, data from row 5) ----
    ws = wb["Fleet Register"]
    order = []
    vehicles = {}
    for r in range(5, ws.max_row + 1):
        reg = ws.cell(row=r, column=4).value
        vid = ws.cell(row=r, column=1).value
        if not reg or not vid:
            continue
        order.append(reg)
        vehicles[reg] = {
            "id": vid,
            "division": ws.cell(row=r, column=2).value or "Unassigned",
            "vehicle": ws.cell(row=r, column=3).value or "",
            "reg": reg,
            "driver": ws.cell(row=r, column=5).value or "Unassigned",
            "type": ws.cell(row=r, column=6).value or "",
            "tracked": ws.cell(row=r, column=7).value or "No",
            "limit": ws.cell(row=r, column=8).value,
            "notes": ws.cell(row=r, column=9).value,
        }

    # ---- Derived: months list (union of fuel + expense months, sorted) ----
    months = sorted(set(f["m"] for f in fuel if f["m"]) | set(e["month"] for e in expenses if e["month"]))

    # ---- Derived: monthlyLitres / monthlyCost per vehicle per month (fuel-sourced) ----
    monthlyLitres = {reg: {m: 0 for m in months} for reg in order}
    monthlyCost = {reg: {m: 0 for m in months} for reg in order}
    for f in fuel:
        if f["r"] in monthlyLitres and f["m"] in monthlyLitres[f["r"]]:
            monthlyLitres[f["r"]][f["m"]] += f["l"]
            monthlyCost[f["r"]][f["m"]] += f["c"]

    # ---- Derived: monthlyOther per vehicle per month (expenses-sourced) ----
    monthlyOther = {reg: {m: 0 for m in months} for reg in order}
    for e in expenses:
        if e["reg"] in monthlyOther and e["month"] in monthlyOther[e["reg"]]:
            monthlyOther[e["reg"]][e["month"]] += e["cost"]

    # ---- Derived: expenseTypeTotals ----
    expenseTypeTotals = defaultdict(float)
    for e in expenses:
        expenseTypeTotals[e["type"]] += e["cost"]

    # ---- Derived: per-vehicle all-time totals + km (odometer delta) + cost/km ----
    for reg in order:
        v = vehicles[reg]
        v["allTimeLitres"] = round(sum(monthlyLitres[reg].values()), 2)
        v["allTimeFuelCost"] = round(sum(monthlyCost[reg].values()), 2)
        v["allTimeOtherCost"] = round(sum(monthlyOther[reg].values()), 2)
        odos = [f["o"] for f in fuel if f["r"] == reg and f["o"] is not None]
        km = (max(odos) - min(odos)) if len(odos) >= 2 else 0
        v["km"] = km
        totalCost = v["allTimeFuelCost"] + v["allTimeOtherCost"]
        v["costPerKm"] = round(totalCost / km, 4) if km > 0 else 0

    out = {
        "vehicles": vehicles,
        "order": order,
        "months": months,
        "monthlyLitres": monthlyLitres,
        "monthlyCost": monthlyCost,
        "monthlyOther": monthlyOther,
        "expenses": expenses,
        "fuel": fuel,
        "expenseTypeTotals": dict(expenseTypeTotals),
    }

    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, separators=(",", ":"))

    print(f"Vehicles: {len(order)}")
    print(f"Months: {months[0]} .. {months[-1]} ({len(months)} months)")
    print(f"Fuel transactions: {len(fuel)}")
    print(f"Expenses: {len(expenses)}")
    print(f"Total fuel cost: N${sum(f['c'] for f in fuel):,.2f}")
    print(f"Total fuel litres: {sum(f['l'] for f in fuel):,.1f} L")
    print(f"Total other expenses: N${sum(e['cost'] for e in expenses):,.2f}")
    print(f"Wrote {out_path}")

if __name__ == "__main__":
    main()
